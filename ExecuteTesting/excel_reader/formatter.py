from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from excel_reader.models import TestCase
from utilities.file_utils import backup_file
from utilities.logger import get_logger

log = get_logger(test="ExcelFormatter")

_RESULT_COLS = ["Status", "Actual Result", "Error", "Screenshot", "Execution Time (ms)", "Failure Analysis"]

_FILLS = {
    "PASS":    PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "FAIL":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "SKIP":    PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "NOT RUN": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _get_or_create_result_cols(ws) -> dict[str, int]:
    """Return column index (1-based) for each result column, creating if missing."""
    max_col = ws.max_column or 1
    header_row = 1
    existing = {ws.cell(header_row, c).value: c for c in range(1, max_col + 1)}

    col_map = {}
    for col_name in _RESULT_COLS:
        if col_name in existing:
            col_map[col_name] = existing[col_name]
        else:
            max_col += 1
            cell = ws.cell(header_row, max_col)
            cell.value = col_name
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(max_col)].width = 20
            col_map[col_name] = max_col

    return col_map


def write_results(file_path: str, test_cases: List[TestCase]) -> None:
    """Append/overwrite result columns in the original Excel file."""
    path = Path(file_path)
    if not path.exists():
        log.error(f"File not found: {file_path}")
        return

    backup_file(file_path)
    log.info(f"Writing results to: {path.name}")

    wb = openpyxl.load_workbook(str(path))

    # Group test cases by sheet
    by_sheet: dict[str, List[TestCase]] = {}
    for tc in test_cases:
        if tc.file_path == file_path:
            by_sheet.setdefault(tc.sheet_name, []).append(tc)

    for sheet_name, cases in by_sheet.items():
        if sheet_name not in wb.sheetnames:
            log.warning(f"Sheet '{sheet_name}' not found in workbook, skipping")
            continue

        ws = wb[sheet_name]
        col_map = _get_or_create_result_cols(ws)

        for tc in cases:
            row = tc.row_index
            fill = _FILLS.get(tc.status, _FILLS["NOT RUN"])

            def write_cell(col_name: str, value: str):
                c = ws.cell(row, col_map[col_name])
                c.value = value
                c.fill = fill
                c.alignment = Alignment(wrap_text=True)

            write_cell("Status", tc.status)
            write_cell("Actual Result", tc.actual_result)
            write_cell("Error", tc.error_message)
            write_cell("Screenshot", tc.screenshot_path)
            write_cell("Execution Time (ms)", str(tc.execution_time_ms))
            write_cell("Failure Analysis", tc.failure_analysis)

        log.info(f"  Updated {len(cases)} rows in sheet '{sheet_name}'")

    wb.save(str(path))
    wb.close()
    log.info(f"Saved: {path.name}")
