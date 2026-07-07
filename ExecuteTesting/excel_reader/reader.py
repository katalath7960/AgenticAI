from pathlib import Path
from typing import List, Dict

import openpyxl

from excel_reader.models import TestCase
from utilities.logger import get_logger

log = get_logger(test="ExcelReader")

# Accepted column name aliases (case-insensitive)
_COL_ALIASES: Dict[str, List[str]] = {
    "tc_id":            ["tc_id", "id", "test id", "test case id", "case id", "no", "#"],
    "test_case_name":   ["test case name", "test case", "name", "title", "scenario"],
    "preconditions":    ["preconditions", "precondition", "pre-conditions", "prerequisites"],
    "steps":            ["steps", "step", "action", "actions", "test steps", "description"],
    "expected_result":  ["expected result", "expected", "expected outcome", "expected results"],
}


def _find_col(headers: List[str], aliases: List[str]) -> int | None:
    for i, h in enumerate(headers):
        if h.strip().lower() in aliases:
            return i
    return None


def _cell_value(cell) -> str:
    val = cell.value
    if val is None:
        return ""
    return str(val).strip()


def read_excel_folder(folder: str) -> List[TestCase]:
    test_cases: List[TestCase] = []
    folder_path = Path(folder)
    xlsx_files = list(folder_path.glob("*.xlsx"))
    if not xlsx_files:
        log.warning(f"No .xlsx files found in '{folder}'")
        return test_cases

    for file_path in sorted(xlsx_files):
        log.info(f"Reading file: {file_path.name}")
        try:
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
        except Exception as e:
            log.error(f"Cannot open {file_path.name}: {e}")
            continue

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=False))
            if not rows:
                continue

            # Detect header row (first non-empty row)
            header_row_idx = 0
            headers = [_cell_value(c) for c in rows[header_row_idx]]
            if not any(headers):
                log.warning(f"  Sheet '{sheet_name}': empty header row, skipping")
                continue

            col_map = {key: _find_col(headers, aliases) for key, aliases in _COL_ALIASES.items()}
            if col_map["steps"] is None:
                log.warning(f"  Sheet '{sheet_name}': no 'Steps' column found, skipping")
                continue

            log.info(f"  Sheet '{sheet_name}': {ws.max_row - 1} data rows")
            for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
                steps_val = _cell_value(row[col_map["steps"]]) if col_map["steps"] is not None else ""
                if not steps_val:
                    continue  # skip blank rows

                def get(key):
                    idx = col_map.get(key)
                    if idx is None:
                        return ""
                    return _cell_value(row[idx])

                auto_id = f"TC_{row_idx - 1:04d}"
                tc = TestCase(
                    file_path=str(file_path),
                    sheet_name=sheet_name,
                    row_index=row_idx,
                    tc_id=get("tc_id") or auto_id,
                    test_case_name=get("test_case_name") or steps_val[:60],
                    preconditions=get("preconditions"),
                    steps=steps_val,
                    expected_result=get("expected_result"),
                )
                test_cases.append(tc)

        wb.close()

    log.info(f"Total test cases loaded: {len(test_cases)}")
    return test_cases
