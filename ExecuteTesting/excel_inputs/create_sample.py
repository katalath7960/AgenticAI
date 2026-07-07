"""Run once to generate the sample test case Excel file."""
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

rows = [
    ["TC_ID", "Test Case Name", "Preconditions", "Steps", "Expected Result"],
    ["TC_001", "Verify Login", "Application is accessible", "Login with valid username and password", "User is logged in successfully and dashboard is displayed"],
    ["TC_002", "Verify Page Title", "User is logged in", "Validate page title", "Page title contains 'Forum'"],
    ["TC_003", "Navigate to Case Search", "User is logged in", "Click menu Case Search", "Case Search page is displayed"],
    ["TC_004", "Search for a Case", "User is on Case Search page", "Enter text 'TEST' in the search field and click Search button", "Search results are displayed"],
    ["TC_005", "Validate Mandatory Fields on Add Case", "User is logged in", "Click button Add Case then click Save without filling mandatory fields", "Error messages are displayed for mandatory fields"],
    ["TC_006", "Add New Case", "User is on Add Case page", "Enter text in all required fields then click Save", "Case is saved successfully and confirmation message is displayed"],
    ["TC_007", "Validate Navigation", "User is logged in", "Validate navigation to Case List page", "Case List page loads correctly"],
    ["TC_008", "Logout", "User is logged in", "Logout from the application", "User is logged out and redirected to login page"],
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Login and Navigation"

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for r_idx, row in enumerate(rows, start=1):
    for c_idx, val in enumerate(row, start=1):
        cell = ws.cell(r_idx, c_idx, val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font

col_widths = [10, 30, 30, 60, 50]
for i, w in enumerate(col_widths, start=1):
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 20
for r in range(2, len(rows) + 1):
    ws.row_dimensions[r].height = 40

out = Path(__file__).parent / "sample_test_cases.xlsx"
wb.save(str(out))
print(f"Created: {out}")
