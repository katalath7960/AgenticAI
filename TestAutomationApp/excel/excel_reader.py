"""Read test cases from an uploaded Excel workbook."""
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Dict, Optional
import openpyxl


@dataclass
class TestCase:
    file_path: str
    sheet_name: str
    row_index: int
    tc_id: str
    test_case_name: str
    preconditions: str
    steps: str
    expected_result: str
    # Populated after execution
    status: str = "NOT RUN"
    actual_result: str = ""
    error_details: str = ""
    screenshot: str = ""
    execution_time: str = ""
    executed_date: str = ""


_COL_ALIASES: Dict[str, List[str]] = {
    "tc_id":           ["tc_id", "id", "test id", "test case id", "case id", "no", "#", "sr no", "sr.no"],
    "test_case_name":  ["test case name", "test case", "name", "title", "scenario", "description", "test name"],
    "preconditions":   ["preconditions", "precondition", "pre-conditions", "prerequisites", "pre conditions"],
    "steps":           ["steps", "step", "action", "actions", "test steps", "test step", "procedure"],
    "expected_result": ["expected result", "expected", "expected outcome", "expected results", "expected behavior"],
}


def _find_col(headers: List[str], aliases: List[str]) -> Optional[int]:
    for i, h in enumerate(headers):
        if str(h).strip().lower() in aliases:
            return i
    return None


def _cell_val(cell) -> str:
    v = cell.value
    return "" if v is None else str(v).strip()


def read_workbook(file_path: str) -> List[TestCase]:
    test_cases: List[TestCase] = []
    path = Path(file_path)

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True)
    except Exception as e:
        raise ValueError(f"Cannot open workbook: {e}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=False))
        if not rows:
            continue

        # Find header row (first non-empty row)
        headers = [_cell_val(c) for c in rows[0]]
        if not any(headers):
            continue

        col_map = {k: _find_col(headers, v) for k, v in _COL_ALIASES.items()}

        if col_map["steps"] is None:
            # Try second row as header if first is empty
            if len(rows) > 1:
                headers = [_cell_val(c) for c in rows[1]]
                col_map = {k: _find_col(headers, v) for k, v in _COL_ALIASES.items()}
            if col_map["steps"] is None:
                continue

        def get(row, key: str) -> str:
            idx = col_map.get(key)
            return _cell_val(row[idx]) if idx is not None else ""

        for row_idx, row in enumerate(rows[1:], start=2):
            step_val = get(row, "steps")
            if not step_val:
                continue
            auto_id = f"TC_{row_idx - 1:04d}"
            tc = TestCase(
                file_path=str(path),
                sheet_name=sheet_name,
                row_index=row_idx,
                tc_id=get(row, "tc_id") or auto_id,
                test_case_name=get(row, "test_case_name") or step_val[:60],
                preconditions=get(row, "preconditions"),
                steps=step_val,
                expected_result=get(row, "expected_result"),
            )
            test_cases.append(tc)

    wb.close()
    return test_cases


def get_sheet_summary(file_path: str) -> Dict[str, int]:
    """Return {sheet_name: row_count} for preview."""
    try:
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        summary = {name: wb[name].max_row - 1 for name in wb.sheetnames}
        wb.close()
        return summary
    except Exception:
        return {}
