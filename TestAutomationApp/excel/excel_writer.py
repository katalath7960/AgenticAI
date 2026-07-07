"""Write execution results back into the original workbook in-place."""
import shutil
from datetime import datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from excel.excel_reader import TestCase

_RESULT_COLS = [
    "Execution Status",
    "Actual Result",
    "Error Details",
    "Screenshot",
    "Execution Time",
    "Executed Date",
]

_FILLS = {
    "PASS":    PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "FAIL":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "SKIP":    PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "NOT RUN": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}

_HDR_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)


def _ensure_result_cols(ws) -> dict[str, int]:
    max_col = ws.max_column or 1
    header_row = 1
    existing = {}
    for c in range(1, max_col + 1):
        val = ws.cell(header_row, c).value
        if val:
            existing[str(val).strip()] = c

    col_map: dict[str, int] = {}
    for col_name in _RESULT_COLS:
        if col_name in existing:
            col_map[col_name] = existing[col_name]
        else:
            max_col += 1
            cell = ws.cell(header_row, max_col)
            cell.value = col_name
            cell.fill = _HDR_FILL
            cell.font = _HDR_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(max_col)].width = 22
            col_map[col_name] = max_col
    return col_map


def write_results(file_path: str, test_cases: List[TestCase]) -> str:
    """
    Write results into the workbook at file_path.
    Creates a backup first. Returns the path to the updated file.
    """
    path = Path(file_path)
    # Backup
    backup = path.with_name(f"{path.stem}_backup{path.suffix}")
    shutil.copy2(path, backup)

    wb = openpyxl.load_workbook(str(path))
    by_sheet: dict[str, List[TestCase]] = {}
    for tc in test_cases:
        by_sheet.setdefault(tc.sheet_name, []).append(tc)

    for sheet_name, cases in by_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        col_map = _ensure_result_cols(ws)

        for tc in cases:
            row = tc.row_index
            fill = _FILLS.get(tc.status, _FILLS["NOT RUN"])

            data = {
                "Execution Status": tc.status,
                "Actual Result":    tc.actual_result,
                "Error Details":    tc.error_details,
                "Screenshot":       tc.screenshot,
                "Execution Time":   tc.execution_time,
                "Executed Date":    tc.executed_date or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            for col_name, value in data.items():
                cell = ws.cell(row, col_map[col_name])
                cell.value = value
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(str(path))
    wb.close()
    # Remove backup after successful save
    backup.unlink(missing_ok=True)
    return str(path)
