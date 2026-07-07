"""
Run once to generate sample_test_cases.xlsx
Usage: python create_sample.py
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

ROWS = [
    # headers
    ["TC_ID", "Test Case Name", "Preconditions", "Steps", "Expected Result"],
    # data
    [
        "TC_001", "Verify Successful Login",
        "Application is accessible at the configured URL",
        "Login with valid username and password",
        "User is successfully logged in and directed to the home/dashboard page",
    ],
    [
        "TC_002", "Verify Page Title",
        "User is logged in",
        "Validate page title",
        "Page title is displayed correctly",
    ],
    [
        "TC_003", "Navigate to Case Search",
        "User is logged in",
        "Click menu Case Search",
        "Case Search page is displayed successfully",
    ],
    [
        "TC_004", "Search for a Case",
        "User is on Case Search page",
        "Enter text 'TEST' in the search field\nClick the Search button",
        "Search results are displayed in the results table",
    ],
    [
        "TC_005", "Validate Search Results Table",
        "Search has been performed",
        "Validate table",
        "Results table is displayed with case records",
    ],
    [
        "TC_006", "Validate Mandatory Field Errors on Add Case",
        "User is logged in",
        "Click menu Add Case\nClick the Save button",
        "Validation error messages are displayed for mandatory fields",
    ],
    [
        "TC_007", "Validate URL after Login",
        "User has just logged in",
        "Validate URL contains 'forum'",
        "URL contains the expected application path",
    ],
    [
        "TC_008", "Validate Error Message",
        "User submitted form with invalid data",
        "Validate error message is displayed",
        "An error message is visible on the page",
    ],
    [
        "TC_009", "Navigate to a Specific Section",
        "User is logged in",
        "Click the Home link",
        "Home page is displayed",
    ],
    [
        "TC_010", "Verify Logout",
        "User is logged in",
        "Logout from the application",
        "User is successfully logged out and redirected to the login page",
    ],
]

HDR_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Regression Suite"

COL_WIDTHS = [10, 35, 35, 60, 55]

for r_idx, row in enumerate(ROWS, start=1):
    for c_idx, val in enumerate(row, start=1):
        cell = ws.cell(r_idx, c_idx, val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if r_idx == 1:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
        elif r_idx % 2 == 0:
            cell.fill = ALT_FILL

for i, w in enumerate(COL_WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 20
for r in range(2, len(ROWS) + 1):
    ws.row_dimensions[r].height = 45

ws.freeze_panes = "A2"

out = Path(__file__).parent / "sample_test_cases.xlsx"
wb.save(str(out))
print(f"Created: {out}")
